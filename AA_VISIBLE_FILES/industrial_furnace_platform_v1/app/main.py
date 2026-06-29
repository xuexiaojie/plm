import base64
import binascii
import logging
import mimetypes
import json
import re
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from time import perf_counter
from uuid import uuid4
from xml.etree import ElementTree

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError
from pypdf import PdfReader
from sqlalchemy.orm import Session
from sqlalchemy.orm import joinedload
import xlrd

from app import models
from app.ai_client import run_joint_analysis
from app.db import SessionLocal, get_db, init_db
from app.executors import run_template
from app.lightrag_retrieval import search_with_lightrag
from app.schemas import (
    ApprovalActionRequest,
    AiAnalysisRequest,
    ArtifactBatchCreate,
    ArtifactCreate,
    ClipboardImageInput,
    ExecutionRequest,
    ExecutorResponse,
    PermissionAssignment,
    ProjectCreate,
    ProjectManagerCandidateCreate,
    ProjectItemCreate,
    ProjectManagementBatchCreate,
    ProjectManagementCreate,
    TemplateCreate,
)
from app.seed import seed_all


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR.parent / "uploaded_artifacts"
INDEX_HTML_PATH = BASE_DIR / "static" / "index.html"
_cached_index_html = ""
_cached_index_mtime_ns = 0
logger = logging.getLogger(__name__)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _build_index_html() -> str:
    with open(INDEX_HTML_PATH, encoding="utf-8") as page:
        html = page.read()
        html = html.replace('id="artifactFileSummary" class="meta">可一次选择多个文档、图片或视频。', 'id="artifactFileSummary" class="meta">可一次选择多个文档、图片或视频。.docx、.xls、.xlsx、PDF 和文本类附件会自动读取正文，图片会提取元信息。')
        html = html.replace('id="artifactBatchDocSummary" class="meta">只支持文档文件，每个文档将生成一个资料条目。', 'id="artifactBatchDocSummary" class="meta">只支持文档文件，每个文档将生成一个资料条目。.docx、.xls、.xlsx、PDF 和文本类附件会自动读取正文。')
        return html


def _get_index_html() -> str:
    global _cached_index_html, _cached_index_mtime_ns
    current_mtime_ns = INDEX_HTML_PATH.stat().st_mtime_ns
    if not _cached_index_html or current_mtime_ns != _cached_index_mtime_ns:
        _cached_index_html = _build_index_html()
        _cached_index_mtime_ns = current_mtime_ns
    return _cached_index_html


@asynccontextmanager
async def lifespan(app: FastAPI):
    global _cached_index_html, _cached_index_mtime_ns
    init_db()
    _cached_index_html = _build_index_html()
    _cached_index_mtime_ns = INDEX_HTML_PATH.stat().st_mtime_ns
    db = SessionLocal()
    try:
        _clear_legacy_artifact_retrieval_state(db)
    finally:
        db.close()
    yield


app = FastAPI(title="工业炉计算平台 V1", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")


ROLE_PERMISSIONS = {
    "engineer": {"execution:run", "approval:submit", "report:create", "report:download", "comparison:create", "artifact:manage", "ai:analyze", "read"},
    "reviewer": {"approval:review", "report:download", "read"},
    "chief_reviewer": {"approval:review", "report:publish", "report:download", "read"},
    "report_admin": {"report:create", "report:publish", "report:download", "read"},
    "template_admin": {"template:manage", "report:download", "read"},
    "algorithm_admin": {"execution:run", "template:manage", "report:download", "read"},
    "admin": {"*", "permission:assign"},
    "readonly": {"read", "report:download"},
}

PERMISSION_CATALOG = [
    {"code": "read", "name": "查看数据", "description": "查看项目、计算、报告和资料数据"},
    {"code": "execution:run", "name": "执行计算", "description": "发起计算执行并生成结果快照"},
    {"code": "approval:submit", "name": "提交审批", "description": "把计算结果提交给审核人"},
    {"code": "approval:review", "name": "审批处理", "description": "执行审批通过或退回"},
    {"code": "report:create", "name": "生成草稿报告", "description": "基于计算结果生成草稿报告"},
    {"code": "report:publish", "name": "发布正式报告", "description": "在审批完成后发布正式报告"},
    {"code": "report:download", "name": "下载报告", "description": "下载已生成报告文本"},
    {"code": "comparison:create", "name": "横向对比", "description": "创建和查看横向对比组"},
    {"code": "artifact:manage", "name": "资料管理", "description": "新增和批量录入项目资料"},
    {"code": "ai:analyze", "name": "AI 分析", "description": "执行 AI 查询和 AI 智能分析"},
    {"code": "template:manage", "name": "模板管理", "description": "维护计算模板和算法入口"},
    {"code": "permission:assign", "name": "权限分配", "description": "维护角色权限配置"},
]

ROLE_NAMES = {
    "engineer": "普通计算人员",
    "reviewer": "专业校核人",
    "chief_reviewer": "总审人",
    "report_admin": "文控专员",
    "template_admin": "模板管理员",
    "algorithm_admin": "算法维护人员",
    "admin": "系统管理员",
    "readonly": "只读用户",
}

BUSINESS_DEPARTMENTS = {"工业炉", "炼铁", "炼钢", "轧钢", "热力", "炉窑设计"}
MANAGEMENT_DEPARTMENTS = {"总部", "总工办", "质量管理部", "科技管理部", "技术研究院", "技术中心", "项目管理", "工艺审核", "结构审核"}
SYSTEM_DEPARTMENTS = {"平台管理"}
DEPARTMENT_ALIASES = {"炉窑设计": "工业炉"}

DEPARTMENT_MODULES = {
    "工业炉": ["flow-analysis-query", "engineering-analysis", "ai-query-view", "artifact-entry-view", "artifact-query-view", "calc-item-management", "approval"],
    "炉窑设计": ["flow-analysis-query", "engineering-analysis", "ai-query-view", "artifact-entry-view", "artifact-query-view", "calc-item-management", "approval"],
    "炼铁": ["flow-analysis-query", "engineering-analysis", "ai-query-view", "artifact-query-view"],
    "炼钢": ["flow-analysis-query", "engineering-analysis", "ai-query-view", "artifact-query-view"],
    "轧钢": ["flow-analysis-query", "engineering-analysis", "ai-query-view", "artifact-query-view"],
    "热力": ["flow-analysis-query", "engineering-analysis", "ai-query-view", "artifact-query-view"],
}

ALL_MENU_MODULES = [
    "project-management",
    "artifact-entry-view",
    "artifact-query-view",
    "calc-item-management",
    "approval",
    "digital-twin",
    "flow-analysis-query",
    "engineering-analysis",
    "permission-view",
]

USER_CATALOG = [
    {"id": 1, "name": "赵总", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 2, "name": "呼启同", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 3, "name": "郭广明", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 4, "name": "吴永红", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 5, "name": "杨小兵", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 6, "name": "梁炜", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 7, "name": "傅巍", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 8, "name": "孟显亮", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 9, "name": "张刚", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 10, "name": "冯威", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 13, "name": "江华", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 14, "name": "朱小辉", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 15, "name": "刘和荣", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 16, "name": "赵云飞", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 17, "name": "杨三堂", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 18, "name": "曹开明", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 19, "name": "王志斌", "role": "admin", "title": "系统管理员", "department": "平台管理"},
    {"id": 81, "name": "工业炉业务员", "role": "engineer", "title": "业务工程师", "department": "工业炉"},
    {"id": 85, "name": "吴启明", "role": "engineer", "title": "业务工程师", "department": "工业炉"},
    {"id": 11, "name": "张工", "role": "engineer", "title": "工艺工程师", "department": "炉窑设计"},
    {"id": 12, "name": "李工", "role": "engineer", "title": "设备工程师", "department": "炉窑设计"},
    {"id": 82, "name": "炼钢业务员", "role": "engineer", "title": "业务工程师", "department": "炼钢"},
    {"id": 83, "name": "总部管理", "role": "readonly", "title": "总部管理人员", "department": "总部"},
    {"id": 84, "name": "科技管理", "role": "readonly", "title": "职能管理人员", "department": "科技管理部"},
    {"id": 21, "name": "王工", "role": "reviewer", "title": "专业校核", "department": "工艺审核"},
    {"id": 22, "name": "周工", "role": "reviewer", "title": "专业校核", "department": "结构审核"},
    {"id": 31, "name": "陈总工", "role": "chief_reviewer", "title": "总审", "department": "技术中心"},
    {"id": 41, "name": "孙文控", "role": "report_admin", "title": "文控专员", "department": "项目管理"},
    {"id": 51, "name": "模板管理员", "role": "template_admin", "title": "模板管理员", "department": "平台管理"},
    {"id": 61, "name": "算法管理员", "role": "algorithm_admin", "title": "算法维护人员", "department": "平台管理"},
    {"id": 71, "name": "访客", "role": "readonly", "title": "只读用户", "department": "访客"},
]
USER_BY_ID = {user["id"]: user for user in USER_CATALOG}
ROLE_USER_IDS = {}
for user in USER_CATALOG:
    ROLE_USER_IDS.setdefault(user["role"], user["id"])
APPROVAL_FLOW_USER_IDS = [21, 31]


ARTIFACT_TYPES = {
    "site_feedback": "现场反馈",
    "drawing_review": "审图单",
    "technical_description": "技术说明",
    "drawing_catalog": "图纸目录",
    "material_list": "材料表",
    "patent_technical_document": "专利等技术文档",
}

PROJECT_MANAGER_CANDIDATES = ["张工", "李工", "王工", "赵工"]
ENTERPRISE_CANDIDATES = ["宝山钢铁股份有限公司", "鞍钢集团工程技术有限公司", "首钢集团有限公司", "河钢集团有限公司"]
TEXT_FILE_SUFFIXES = {".txt", ".md", ".csv", ".json", ".html", ".xml", ".yaml", ".yml", ".log"}
EXCEL_FILE_SUFFIXES = {".xls", ".xlsx"}
IMAGE_FILE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}
PASTED_IMAGE_SUMMARY_LIMIT = 500


@app.get("/", response_class=HTMLResponse)
def index() -> HTMLResponse:
    return HTMLResponse(
        content=_get_index_html(),
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


def require_permission(permission: str, role: str | None = Header(default="admin", alias="X-Role")) -> str:
    permissions = ROLE_PERMISSIONS.get(role or "", set())
    if "*" not in permissions and permission not in permissions:
        raise HTTPException(status_code=403, detail={"code": "PERMISSION_DENIED", "message": "权限不足"})
    return role or "admin"


def permission_dependency(permission: str):
    def dependency(x_role: str | None = Header(default="admin", alias="X-Role")) -> str:
        return require_permission(permission, x_role)

    return dependency


def get_current_user(
    x_role: str | None = Header(default="admin", alias="X-Role"),
    x_user_id: int | None = Header(default=None, alias="X-User-Id"),
) -> dict:
    if x_user_id is not None:
        user = USER_BY_ID.get(x_user_id)
        if user is None:
            raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "当前用户不存在"})
        if x_role and x_role != user["role"]:
            raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "用户与角色不匹配"})
        return user
    role_code = x_role or "admin"
    user_id = ROLE_USER_IDS.get(role_code, ROLE_USER_IDS["admin"])
    return USER_BY_ID[user_id]


def get_current_user_id(current_user: dict = Depends(get_current_user)) -> int:
    return int(current_user["id"])


def department_access_scope(user: dict) -> dict:
    department = str(user.get("department") or "")
    access_department = DEPARTMENT_ALIASES.get(department, department)
    role = str(user.get("role") or "")
    if role == "admin" or department in SYSTEM_DEPARTMENTS:
        level = "system"
        allowed_departments = sorted(BUSINESS_DEPARTMENTS | MANAGEMENT_DEPARTMENTS | SYSTEM_DEPARTMENTS | {department, access_department})
        visible_modules = ALL_MENU_MODULES
    elif department in MANAGEMENT_DEPARTMENTS:
        level = "management"
        allowed_departments = sorted(BUSINESS_DEPARTMENTS | {department, access_department})
        visible_modules = ALL_MENU_MODULES
    else:
        level = "department"
        allowed_departments = sorted({department, access_department})
        visible_modules = ["project-management", *DEPARTMENT_MODULES.get(access_department, DEPARTMENT_MODULES["工业炉"])]
    return {
        "level": level,
        "department": department,
        "access_department": access_department,
        "allowed_departments": allowed_departments,
        "visible_modules": visible_modules,
        "physical_storage_root": f"uploaded_artifacts/{department}",
    }


def _can_access_project(project: models.Project | None, current_user: dict) -> bool:
    if project is None or project.deleted_at is not None:
        return False
    scope = department_access_scope(current_user)
    return scope["level"] in {"system", "management"} or project.department in scope["allowed_departments"]


def require_project_access(db: Session, project_id: int, current_user: dict) -> models.Project:
    project = db.get(models.Project, project_id)
    if not _can_access_project(project, current_user):
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "项目不存在或无权访问"})
    return project


def visible_project_query(db: Session, current_user: dict):
    query = db.query(models.Project).filter(models.Project.deleted_at.is_(None))
    scope = department_access_scope(current_user)
    if scope["level"] not in {"system", "management"}:
        query = query.filter(models.Project.department.in_(scope["allowed_departments"]))
    return query


def _project_payload(project: models.Project) -> dict:
    return {"id": project.id, "code": project.code, "name": project.name, "department": project.department, "status": project.status}


def _department_from_payload(payload_department: str | None, current_user: dict) -> str:
    scope = department_access_scope(current_user)
    requested = (payload_department or "工业炉").strip()
    if scope["level"] == "department" and requested not in scope["allowed_departments"]:
        raise HTTPException(status_code=403, detail={"code": "PERMISSION_DENIED", "message": "普通业务部门只能创建本部门项目"})
    return requested


def _permission_snapshot() -> dict:
    catalog_codes = {item["code"] for item in PERMISSION_CATALOG}
    role_codes = sorted({permission for permissions in ROLE_PERMISSIONS.values() for permission in permissions if permission != "*"})
    missing_definitions = [permission for permission in role_codes if permission not in catalog_codes]
    unused_definitions = sorted(catalog_codes - set(role_codes))
    return {
        "permissions": PERMISSION_CATALOG,
        "roles": [
            {
                "code": role,
                "name": ROLE_NAMES.get(role, role),
                "permissions": sorted(permissions),
                "is_super_admin": "*" in permissions,
            }
            for role, permissions in ROLE_PERMISSIONS.items()
        ],
        "self_check": {
            "ok": not missing_definitions,
            "missing_definitions": missing_definitions,
            "unused_definitions": unused_definitions,
        },
    }


def validate_project_item(db: Session, project_id: int, project_item_id: int | None) -> None:
    if project_item_id is None:
        return
    item = db.get(models.ProjectItem, project_item_id)
    if item is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "名目不存在"})
    if item.project_id != project_id:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "名目不属于当前项目"})


def _extract_docx_text(content: bytes) -> str:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            document = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile) as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "Word 文档解析失败，请确认文件为 .docx 格式"}) from error
    root = ElementTree.fromstring(document)
    paragraphs = []
    for paragraph in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
        text = "".join(node.text or "" for node in paragraph.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"))
        if text.strip():
            paragraphs.append(text.strip())
    return "\n".join(paragraphs)


def _extract_pdf_text(content: bytes) -> str:
    try:
        reader = PdfReader(BytesIO(content))
        pages = [page.extract_text() or "" for page in reader.pages]
    except Exception as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "PDF 文档解析失败，请确认文件为有效 PDF 格式"}) from error
    return "\n".join(page.strip() for page in pages if page.strip())


def _extract_xlsx_text(content: bytes) -> str:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "Excel 文档解析失败，请确认文件为有效 .xlsx 格式"}) from error
    rows = []
    for sheet in workbook.worksheets:
        rows.append(f"工作表: {sheet.title}")
        for values in sheet.iter_rows(values_only=True):
            cells = [str(value).strip() for value in values if value is not None and str(value).strip()]
            if cells:
                rows.append(" | ".join(cells))
    return "\n".join(rows)


def _extract_xls_text(content: bytes) -> str:
    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "Excel 文档解析失败，请确认文件为有效 .xls 格式"}) from error
    rows = []
    for sheet in workbook.sheets():
        rows.append(f"工作表: {sheet.name}")
        for row_index in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_index, column_index)).strip() for column_index in range(sheet.ncols)]
            cells = [cell for cell in cells if cell]
            if cells:
                rows.append(" | ".join(cells))
    return "\n".join(rows)


def _extract_image_summary(filename: str, content: bytes) -> str:
    if len(content) < 32:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "解析失败：图片数据过短"})
    try:
        with Image.open(BytesIO(content)) as image:
            image.verify()
        with Image.open(BytesIO(content)) as image:
            width, height = image.size
            return "\n".join(
                [
                    f"文件名: {filename}",
                    f"图片格式: {image.format or '未知'}",
                    f"尺寸: {width} x {height}",
                    f"颜色模式: {image.mode or '未知'}",
                    f"帧数: {getattr(image, 'n_frames', 1)}",
                ]
            )
    except UnidentifiedImageError as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "解析失败：非有效图片文件"}) from error
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": f"图片校验失败：{error}"}) from error


def _decode_uploaded_text(filename: str, content_type: str | None, content: bytes) -> tuple[str, str]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".docx":
        return "已解析 Word 正文", _extract_docx_text(content)
    if suffix == ".pdf" or content_type == "application/pdf":
        text = _extract_pdf_text(content)
        return ("已解析 PDF 文本" if text.strip() else "PDF 未提取到可复制文本", text)
    if suffix == ".xlsx":
        text = _extract_xlsx_text(content)
        return ("已解析 Excel 表格" if text.strip() else "Excel 未提取到单元格内容", text)
    if suffix == ".xls":
        text = _extract_xls_text(content)
        return ("已解析 Excel 表格" if text.strip() else "Excel 未提取到单元格内容", text)
    if (content_type or "").startswith("text/") or suffix in TEXT_FILE_SUFFIXES:
        for encoding in ("utf-8", "gb18030"):
            try:
                return "已读取文本正文", content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return "文本附件解码失败", ""
    if (content_type or "").startswith("image/") or suffix in IMAGE_FILE_SUFFIXES:
        return "已提取图片元信息", _extract_image_summary(filename, content)
    return "当前版本支持自动解析 .docx、.xls、.xlsx、PDF 和文本类附件正文，也会提取图片元信息", ""


def _safe_upload_filename(filename: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_.\-\u4e00-\u9fff]+", "_", filename).strip("._")
    return safe or "unnamed"


def _project_storage_dir(db: Session, project_id: int) -> Path:
    project = db.get(models.Project, project_id)
    department = _safe_upload_filename(project.department if project else "unknown")
    return UPLOAD_DIR / department / str(project_id)


def _save_uploaded_file(db: Session, project_id: int, artifact_id: int, filename: str, content: bytes) -> Path:
    directory = _project_storage_dir(db, project_id) / str(artifact_id)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / _safe_upload_filename(filename)
    path.write_bytes(content)
    return path


def _extract_uploaded_filename(content: str) -> str | None:
    match = re.search(r"^- (.+?) \| .+? \|", content, flags=re.MULTILINE)
    return match.group(1).strip() if match else None


def _extract_uploaded_prefix(content: str) -> str:
    marker = "\n\n附件清单与正文:"
    prefix, _, _ = content.partition(marker)
    return prefix.strip()


def _extract_uploaded_content_type(content: str, filename: str) -> str:
    match = re.search(r"^- .+? \| (.+?) \|", content, flags=re.MULTILINE)
    if match and match.group(1).strip():
        return match.group(1).strip()
    guessed, _ = mimetypes.guess_type(filename)
    return guessed or "application/octet-stream"


def _build_artifact_content(prefix: str, filename: str, content_type: str, raw: bytes, status: str, text: str) -> str:
    file_summary = [
        "附件清单与正文:",
        f"- {filename} | {content_type or '未知类型'} | {len(raw) / 1024 / 1024:.2f} MB",
        f"  说明: {status}",
    ]
    if text.strip():
        file_summary.extend(["  正文:", text[:50000]])
    return "\n".join([prefix, "", *file_summary]).strip()


def _decode_data_url_bytes(data_url: str) -> bytes:
    clean_url = re.sub(r"\s+", "", (data_url or "").strip())
    match = re.search(r"data:([^;,]+);base64,([^\"'<>]+)", clean_url, flags=re.IGNORECASE)
    if not match:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "粘贴图片必须为标准 data-url 格式（image/*）"})
    mime_type, raw_base64 = match.groups()
    if not mime_type.lower().startswith("image/"):
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "仅支持粘贴图片，不支持其他文件"})
    normalized_base64 = raw_base64.replace("-", "+").replace("_", "/")
    padding = (4 - len(normalized_base64) % 4) % 4
    normalized_base64 += "=" * padding
    try:
        return base64.b64decode(normalized_base64, validate=True)
    except binascii.Error as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": f"Base64 解码失败：{error}"}) from error
    except Exception as error:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": f"图片二进制解析异常：{error}"}) from error


def _build_pasted_image_context(images: list[ClipboardImageInput]) -> list[dict]:
    rows = []
    for image in images:
        raw = _decode_data_url_bytes(image.data_url)
        status, text = _decode_uploaded_text(image.name, image.content_type, raw)
        rows.append(
            {
                "name": image.name,
                "content_type": image.content_type,
                "parse_status": status,
                "summary": text[:PASTED_IMAGE_SUMMARY_LIMIT],
            }
        )
    return rows


def _soft_delete_artifact(db: Session, artifact: models.ProjectArtifact) -> None:
    artifact.status = "DELETED"
    artifact.deleted_at = utc_now()
    _clear_artifact_chunks(db, artifact_id=artifact.id)


def _soft_delete_matching_uploads(db: Session, project_id: int, artifact_id: int, filename: str) -> int:
    deleted = 0
    for artifact in db.query(models.ProjectArtifact).filter_by(project_id=project_id, status="ACTIVE").all():
        if artifact.id == artifact_id:
            continue
        if _extract_uploaded_filename(artifact.content) == filename:
            _soft_delete_artifact(db, artifact)
            deleted += 1
    return deleted


def _find_stored_upload(project_id: int, artifact_id: int, filename: str, db: Session | None = None) -> Path | None:
    if db is not None:
        path = _project_storage_dir(db, project_id) / str(artifact_id) / _safe_upload_filename(filename)
        if path.exists():
            return path
    path = UPLOAD_DIR / str(project_id) / str(artifact_id) / _safe_upload_filename(filename)
    if path.exists():
        return path
    matches = list(UPLOAD_DIR.glob(f"**/{_safe_upload_filename(filename)}"))
    return matches[0] if matches else None


def _reparse_stored_artifacts(db: Session) -> dict:
    parsed = []
    missing_files = []
    skipped = []
    artifacts = db.query(models.ProjectArtifact).filter_by(status="ACTIVE").all()
    for artifact in artifacts:
        filename = _extract_uploaded_filename(artifact.content)
        if not filename:
            skipped.append({"id": artifact.id, "title": artifact.title, "reason": "未找到已上传文件名"})
            continue
        path = _find_stored_upload(artifact.project_id, artifact.id, filename, db)
        if path is None:
            missing_files.append({"id": artifact.id, "title": artifact.title, "filename": filename})
            continue
        raw = path.read_bytes()
        content_type = _extract_uploaded_content_type(artifact.content, filename)
        status, text = _decode_uploaded_text(filename, content_type, raw)
        prefix = _extract_uploaded_prefix(artifact.content)
        artifact.content = _build_artifact_content(prefix, filename, content_type, raw, status, text)
        _clear_artifact_chunks(db, artifact_id=artifact.id)
        parsed.append({"id": artifact.id, "title": artifact.title, "filename": filename, "parse_status": status, "text_length": len(text)})
    if parsed:
        db.commit()
    return {"parsed_count": len(parsed), "missing_count": len(missing_files), "skipped_count": len(skipped), "parsed": parsed, "missing_files": missing_files, "skipped": skipped}


def _artifact_response(artifact: models.ProjectArtifact) -> dict:
    return {"id": artifact.id, "artifact_type": artifact.artifact_type, "type_name": ARTIFACT_TYPES[artifact.artifact_type], "title": artifact.title}


def _artifact_file_payload(artifact: models.ProjectArtifact, db: Session | None = None) -> dict:
    filename = _extract_uploaded_filename(artifact.content)
    if not filename:
        return {"file_name": None, "file_content_type": None, "has_file": False, "view_url": None}
    path = _find_stored_upload(artifact.project_id, artifact.id, filename, db)
    content_type = _extract_uploaded_content_type(artifact.content, filename)
    return {
        "file_name": filename,
        "file_content_type": content_type,
        "has_file": path is not None,
        "view_url": f"/api/artifacts/{artifact.id}/file" if path is not None else None,
    }


def _artifact_list_payload(artifact: models.ProjectArtifact, db: Session | None = None) -> dict:
    return {
        "id": artifact.id,
        "project_item_id": artifact.project_item_id,
        "artifact_type": artifact.artifact_type,
        "type_name": ARTIFACT_TYPES.get(artifact.artifact_type, artifact.artifact_type),
        "title": artifact.title,
        "source_code": artifact.source_code,
        "content_preview": _content_preview(artifact.content),
        "content_length": len(artifact.content),
        **_artifact_file_payload(artifact, db),
    }


def _content_preview(content: str, limit: int = 240) -> str:
    compact = " ".join(content.split())
    if len(compact) <= limit:
        return compact
    return compact[:limit].rstrip() + "..."


def _clear_artifact_chunks(db: Session, artifact_id: int | None = None) -> int:
    query = db.query(models.ProjectArtifactChunk)
    if artifact_id is not None:
        query = query.filter_by(artifact_id=artifact_id)
    return query.delete()


def _clear_legacy_artifact_retrieval_state(db: Session) -> None:
    removed_chunks = db.query(models.ProjectArtifactChunk).delete()
    removed_analyses = db.query(models.AiAnalysis).filter(models.AiAnalysis.request_json.contains("retrieved_chunks")).delete()
    if removed_chunks or removed_analyses:
        db.commit()


def _tokenize_search_terms(text: str) -> list[str]:
    raw = (text or "").lower()
    tokens = {token for token in re.findall(r"[a-z0-9_]+", raw) if token.strip()}
    for phrase in re.findall(r"[\u4e00-\u9fff]{2,}", raw):
        phrase = phrase.strip()
        if not phrase:
            continue
        tokens.add(phrase)
        max_size = min(4, len(phrase))
        for size in range(max_size, 1, -1):
            for index in range(0, len(phrase) - size + 1):
                tokens.add(phrase[index:index + size])
    if any(term in raw for term in ("方坯尺寸", "坯料尺寸", "钢坯尺寸", "坯料规格", "方坯规格")):
        tokens.update({"坯料断面", "坯料", "断面"})
    if "图号" in raw:
        tokens.update({"doc", "doc no", "doc. no"})
    return sorted(tokens, key=lambda token: (-len(token), token))


def _search_query_variants(question: str) -> list[str]:
    raw = (question or "").strip()
    if not raw:
        return []
    variants = [raw]
    normalized = raw.replace("出路温度", "出炉温度")
    if normalized != raw:
        variants.append(normalized)
    if any(term in normalized for term in ("出炉温度", "入炉温度", "坯料断面", "炉底机械传动", "支撑梁冷却方式", "图号")):
        variants.append(f"{normalized} 技术性能表")
    keywords = _tokenize_search_terms(normalized)
    if keywords:
        variants.append(" ".join(keywords[:6]))
    deduped: list[str] = []
    for variant in variants:
        compact = variant.strip()
        if compact and compact not in deduped:
            deduped.append(compact)
    return deduped


def _artifact_search_text(artifact: models.ProjectArtifact) -> str:
    return "\n".join([artifact.title or "", artifact.source_code or "", artifact.content or ""]).lower()


def _artifact_excerpt(text: str, keywords: list[str], limit: int = 260) -> str:
    compact = re.sub(r"\s+", " ", text or "").strip()
    if not compact:
        return ""
    anchor = 0
    for keyword in keywords:
        index = compact.lower().find(keyword)
        if index >= 0:
            anchor = max(0, index - 40)
            break
    snippet = compact[anchor:anchor + limit]
    return snippet if len(compact) <= anchor + limit else snippet.rstrip() + "..."


def _artifact_ai_content(text: str, keywords: list[str], limit: int = 8000) -> str:
    compact = re.sub(r"\s+", " ", text or "").strip()
    if not compact:
        return ""
    if len(compact) <= limit:
        return compact
    return _artifact_excerpt(compact, keywords, limit=2200)


def _merge_retrieved_rows(row_groups: list[list[dict]], limit: int = 8) -> list[dict]:
    merged: dict[int, dict] = {}
    for row_group in row_groups:
        for row in row_group:
            artifact_id = int(row.get("artifact_id") or 0)
            if artifact_id <= 0:
                continue
            score = int(row.get("score") or 0)
            provider_bonus = 5 if row.get("retrieval_provider") == "lightrag" else 0
            total_score = score + provider_bonus
            existing = merged.get(artifact_id)
            candidate = {**row, "score": total_score}
            if existing is None or total_score > int(existing.get("score") or 0):
                merged[artifact_id] = candidate
            elif total_score == int(existing.get("score") or 0):
                if len(str(row.get("content") or "")) > len(str(existing.get("content") or "")):
                    merged[artifact_id] = candidate
    return sorted(merged.values(), key=lambda row: (int(row.get("score") or 0), int(row.get("artifact_id") or 0)), reverse=True)[:limit]


def _search_project_artifacts(db: Session, project_id: int, question: str, artifact_ids: list[int], limit: int = 8) -> list[dict]:
    query = db.query(models.ProjectArtifact).filter_by(project_id=project_id, status="ACTIVE")
    if artifact_ids:
        query = query.filter(models.ProjectArtifact.id.in_(artifact_ids))
    keywords = _tokenize_search_terms(question)
    complex_question = any(len(keyword) >= 8 or re.search(r"[a-z0-9]", keyword) for keyword in keywords)
    scored = []
    for artifact in query.all():
        haystack = _artifact_search_text(artifact)
        if not haystack.strip():
            continue
        score = 0
        matched_keywords: set[str] = set()
        for keyword in keywords:
            if keyword in haystack:
                score += max(3, len(keyword))
                matched_keywords.add(keyword)
            if keyword in (artifact.title or "").lower():
                score += max(2, len(keyword))
                matched_keywords.add(keyword)
            if keyword in (artifact.source_code or "").lower():
                score += 1
            negative_match = re.search(rf"{re.escape(keyword)}.{{0,8}}(无关|不涉及|未出现|没有|不包含)|(?:无关|不涉及|未出现|没有|不包含).{{0,8}}{re.escape(keyword)}", haystack)
            if negative_match:
                score -= max(3, len(keyword))
        if not keywords:
            score = 1
        max_matched_length = max((len(keyword) for keyword in matched_keywords), default=0)
        if complex_question and max_matched_length < 3:
            continue
        if score <= 0:
            continue
        scored.append((score, artifact, _artifact_excerpt(artifact.content, keywords)))
    if not scored and keywords:
        return []
    if not scored:
        fallback = query.order_by(models.ProjectArtifact.id.desc()).limit(limit).all()
        return [
            {
                "artifact_id": artifact.id,
                "score": 0,
                "type": artifact.artifact_type,
                "type_name": ARTIFACT_TYPES.get(artifact.artifact_type, artifact.artifact_type),
                "title": artifact.title,
                "content": _artifact_ai_content(artifact.content, []),
            }
            for artifact in fallback
        ]
    scored.sort(key=lambda item: (item[0], item[1].id), reverse=True)
    return [
        {
            "artifact_id": artifact.id,
            "score": score,
            "type": artifact.artifact_type,
            "type_name": ARTIFACT_TYPES.get(artifact.artifact_type, artifact.artifact_type),
            "title": artifact.title,
            "content": _artifact_ai_content(artifact.content, keywords),
        }
        for score, artifact, snippet in scored[:limit]
    ]


async def _search_project_artifacts_for_ai(db: Session, project_id: int, question: str, artifact_ids: list[int], limit: int = 8) -> list[dict]:
    query = db.query(models.ProjectArtifact).filter_by(project_id=project_id, status="ACTIVE")
    if artifact_ids:
        query = query.filter(models.ProjectArtifact.id.in_(artifact_ids))
    artifacts = query.order_by(models.ProjectArtifact.id.desc()).all()

    row_groups: list[list[dict]] = []

    try:
        lightrag_rows = await search_with_lightrag(project_id, question, artifacts, limit=limit)
    except Exception:
        logger.exception("LightRAG retrieval failed, falling back to keyword search", extra={"project_id": project_id})
    else:
        if lightrag_rows:
            for row in lightrag_rows:
                row["type_name"] = ARTIFACT_TYPES.get(row.get("type"), row.get("type"))
            row_groups.append(lightrag_rows)

    selected_ids = [artifact.id for artifact in artifacts]
    for variant in _search_query_variants(question):
        keyword_rows = _search_project_artifacts(db, project_id, variant, selected_ids, limit=limit)
        if keyword_rows:
            row_groups.append(keyword_rows)

    merged = _merge_retrieved_rows(row_groups, limit=limit)
    if merged:
        return merged
    return _search_project_artifacts(db, project_id, question, selected_ids, limit=limit)


@app.post("/api/seed")
def seed(db: Session = Depends(get_db)) -> dict[str, str]:
    seed_all(db)
    return {"status": "ok"}


@app.get("/api/projects")
def list_projects(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    projects = visible_project_query(db, current_user).all()
    return [_project_payload(project) for project in projects]


@app.get("/api/project-management/options")
def get_project_management_options() -> dict[str, list[str]]:
    return {"project_managers": PROJECT_MANAGER_CANDIDATES, "enterprises": ENTERPRISE_CANDIDATES}


@app.get("/api/project-managers")
def list_project_managers() -> list[dict]:
    return [{"name": name} for name in PROJECT_MANAGER_CANDIDATES]


@app.post("/api/project-managers")
def create_project_manager(payload: ProjectManagerCandidateCreate) -> dict:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "项目经理姓名不能为空"})
    if name in PROJECT_MANAGER_CANDIDATES:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "项目经理已存在"})
    PROJECT_MANAGER_CANDIDATES.append(name)
    PROJECT_MANAGER_CANDIDATES.sort(key=locale_key)
    return {"name": name, "count": len(PROJECT_MANAGER_CANDIDATES)}


@app.put("/api/project-managers/{manager_name}")
def update_project_manager(manager_name: str, payload: ProjectManagerCandidateCreate) -> dict:
    current_name = manager_name.strip()
    next_name = payload.name.strip()
    if not current_name or not next_name:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "项目经理姓名不能为空"})
    if current_name not in PROJECT_MANAGER_CANDIDATES:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "项目经理不存在"})
    if next_name != current_name and next_name in PROJECT_MANAGER_CANDIDATES:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "项目经理已存在"})
    index = PROJECT_MANAGER_CANDIDATES.index(current_name)
    PROJECT_MANAGER_CANDIDATES[index] = next_name
    PROJECT_MANAGER_CANDIDATES.sort(key=locale_key)
    return {"name": next_name, "count": len(PROJECT_MANAGER_CANDIDATES)}


@app.delete("/api/project-managers/{manager_name}")
def delete_project_manager(manager_name: str) -> dict:
    current_name = manager_name.strip()
    if current_name not in PROJECT_MANAGER_CANDIDATES:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "项目经理不存在"})
    PROJECT_MANAGER_CANDIDATES.remove(current_name)
    return {"name": current_name, "status": "DELETED", "count": len(PROJECT_MANAGER_CANDIDATES)}


@app.get("/api/permissions")
def list_permissions(role: str = Depends(permission_dependency("read"))) -> dict:
    return _permission_snapshot()


@app.get("/api/users")
def list_users(role: str = Depends(permission_dependency("read"))) -> list[dict]:
    return [
        {
            **user,
            "role_name": ROLE_NAMES.get(user["role"], user["role"]),
            "permissions": sorted(ROLE_PERMISSIONS.get(user["role"], set())),
            "access_scope": department_access_scope(user),
        }
        for user in USER_CATALOG
    ]


@app.get("/api/current-user")
def get_current_user_profile(current_user: dict = Depends(get_current_user)) -> dict:
    return {
        **current_user,
        "role_name": ROLE_NAMES.get(current_user["role"], current_user["role"]),
        "permissions": sorted(ROLE_PERMISSIONS.get(current_user["role"], set())),
        "access_scope": department_access_scope(current_user),
    }


@app.put("/api/permissions/roles/{role_code}")
def update_role_permissions(role_code: str, payload: PermissionAssignment, role: str = Depends(permission_dependency("permission:assign"))) -> dict:
    if role_code not in ROLE_PERMISSIONS:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "角色不存在"})
    if role_code == "admin":
        raise HTTPException(status_code=400, detail={"code": "STATE_INVALID", "message": "系统管理员权限固定为全部权限"})
    allowed_permissions = {item["code"] for item in PERMISSION_CATALOG}
    invalid_permissions = sorted(set(payload.permissions) - allowed_permissions)
    if invalid_permissions:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "存在不支持的权限", "details": {"permissions": invalid_permissions}})
    ROLE_PERMISSIONS[role_code] = set(payload.permissions)
    return _permission_snapshot()


@app.post("/api/projects")
def create_project(payload: ProjectCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    existing = db.query(models.Project).filter_by(code=payload.code).one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "项目编码已存在"})
    data = payload.model_dump()
    data["department"] = _department_from_payload(data.get("department"), current_user)
    project = models.Project(**data)
    db.add(project)
    db.commit()
    db.refresh(project)
    return _project_payload(project)


def _project_description(project_manager: str, enterprise: str, technical_terms: str | None, created_at: str | None = None) -> str:
    return json.dumps(
        {"project_manager": project_manager, "enterprise": enterprise, "created_at_input": created_at or "", "technical_terms": technical_terms or ""},
        ensure_ascii=False,
    )


def _project_management_row(project: models.Project) -> dict:
    details = {}
    if project.description:
        try:
            details = json.loads(project.description)
        except json.JSONDecodeError:
            details = {"technical_terms": project.description}
    return {
        "id": project.id,
        "project_name": project.name,
        "department": project.department,
        "project_manager": details.get("project_manager", ""),
        "created_at": details.get("created_at_input") or project.created_at.isoformat(),
        "enterprise": details.get("enterprise", ""),
        "technical_terms": details.get("technical_terms", ""),
    }


def locale_key(value: str) -> str:
    return value.casefold()


def _soft_delete_project(db: Session, project: models.Project) -> None:
    project.deleted_at = utc_now()
    project.status = "DELETED"
    items = db.query(models.ProjectItem).filter_by(project_id=project.id).all()
    for item in items:
        item.deleted_at = utc_now()
        item.status = "DELETED"
    artifacts = db.query(models.ProjectArtifact).filter_by(project_id=project.id, status="ACTIVE").all()
    for artifact in artifacts:
        _soft_delete_artifact(db, artifact)


@app.get("/api/project-management/projects")
def list_project_management_projects(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    projects = visible_project_query(db, current_user).order_by(models.Project.id.desc()).all()
    return [_project_management_row(project) for project in projects]


@app.post("/api/project-management/projects")
def create_project_management_project(payload: ProjectManagementCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    project = models.Project(
        code=f"PRJ-MGMT-{utc_now().strftime('%Y%m%d%H%M%S%f')}-{uuid4().hex[:6]}",
        name=payload.project_name,
        owner_user_id=int(current_user["id"]),
        department=_department_from_payload(payload.department, current_user),
        status="ACTIVE",
        description=_project_description(payload.project_manager, payload.enterprise, payload.technical_terms, payload.created_at),
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    return _project_management_row(project)


@app.put("/api/project-management/projects/{project_id}")
def update_project_management_project(project_id: int, payload: ProjectManagementCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    project = require_project_access(db, project_id, current_user)
    project.name = payload.project_name
    project.description = _project_description(payload.project_manager, payload.enterprise, payload.technical_terms, payload.created_at)
    db.commit()
    db.refresh(project)
    return _project_management_row(project)


@app.delete("/api/project-management/projects/{project_id}")
def delete_project_management_project(project_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    project = require_project_access(db, project_id, current_user)
    _soft_delete_project(db, project)
    db.commit()
    return {"id": project_id, "status": "DELETED"}


@app.post("/api/project-management/projects/batch")
def create_project_management_projects_batch(payload: ProjectManagementBatchCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    if not payload.items:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "批量项目不能为空"})
    projects = []
    for item in payload.items:
        project = models.Project(
            code=f"PRJ-MGMT-{utc_now().strftime('%Y%m%d%H%M%S%f')}-{uuid4().hex[:6]}",
            name=item.project_name,
            owner_user_id=int(current_user["id"]),
            department=_department_from_payload(item.department, current_user),
            status="ACTIVE",
            description=_project_description(item.project_manager, item.enterprise, item.technical_terms, item.created_at),
        )
        db.add(project)
        projects.append(project)
    db.commit()
    for project in projects:
        db.refresh(project)
    return {"count": len(projects), "items": [_project_management_row(project) for project in projects]}


@app.post("/api/projects/{project_id}/items")
def create_item(project_id: int, payload: ProjectItemCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    require_project_access(db, project_id, current_user)
    item = models.ProjectItem(project_id=project_id, **payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id, "code": item.code, "name": item.name}


@app.get("/api/projects/{project_id}/items")
def list_items(project_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    require_project_access(db, project_id, current_user)
    items = db.query(models.ProjectItem).filter_by(project_id=project_id).filter(models.ProjectItem.deleted_at.is_(None)).all()
    return [{"id": i.id, "code": i.code, "name": i.name, "furnace_type": i.furnace_type} for i in items]


@app.get("/api/calc-items")
def list_calc_items(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    visible_project_ids = [project.id for project in visible_project_query(db, current_user).all()]
    items = db.query(models.ProjectItem).filter(models.ProjectItem.deleted_at.is_(None), models.ProjectItem.project_id.in_(visible_project_ids)).order_by(models.ProjectItem.id.desc()).all()
    return [
        {
            "id": item.id,
            "project_id": item.project_id,
            "code": item.code,
            "name": item.name,
            "furnace_type": item.furnace_type,
            "business_scope": item.business_scope,
            "design_stage": item.design_stage,
            "status": item.status,
        }
        for item in items
    ]


@app.delete("/api/calc-items/{item_id}")
def delete_calc_item(item_id: int, db: Session = Depends(get_db)) -> dict:
    item = db.get(models.ProjectItem, item_id)
    if item is None or item.deleted_at is not None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "计算名目不存在"})
    item.deleted_at = utc_now()
    item.status = "DELETED"
    db.commit()
    return {"id": item.id, "status": item.status}


@app.get("/api/artifact-types")
def list_artifact_types() -> list[dict]:
    return [{"code": code, "name": name} for code, name in ARTIFACT_TYPES.items()]


@app.post("/api/projects/{project_id}/artifacts")
def create_artifact(
    project_id: int,
    payload: ArtifactCreate,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("artifact:manage")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    if payload.artifact_type not in ARTIFACT_TYPES:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "资料类型不支持"})
    require_project_access(db, project_id, current_user)
    validate_project_item(db, project_id, payload.project_item_id)
    artifact = models.ProjectArtifact(project_id=project_id, **payload.model_dump())
    db.add(artifact)
    db.flush()
    _clear_artifact_chunks(db, artifact_id=artifact.id)
    db.commit()
    db.refresh(artifact)
    return _artifact_response(artifact)


@app.post("/api/projects/{project_id}/artifacts/upload")
async def upload_artifact(
    project_id: int,
    artifact_type: str = Form(...),
    title: str = Form(...),
    content: str = Form(""),
    source_code: str | None = Form(None),
    project_item_id: int | None = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("artifact:manage")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    if artifact_type not in ARTIFACT_TYPES:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "资料类型不支持"})
    require_project_access(db, project_id, current_user)
    validate_project_item(db, project_id, project_item_id)
    raw = await file.read()
    status, text = _decode_uploaded_text(file.filename or title, file.content_type, raw)
    file_summary = [
        "附件清单与正文:",
        f"- {file.filename or '未命名文件'} | {file.content_type or '未知类型'} | {len(raw) / 1024 / 1024:.2f} MB",
        f"  说明: {status}",
    ]
    if text.strip():
        file_summary.extend(["  正文:", text[:50000]])
    artifact = models.ProjectArtifact(
        project_id=project_id,
        project_item_id=project_item_id,
        artifact_type=artifact_type,
        title=title,
        source_code=source_code,
        content="\n".join([content or "", "", *file_summary]).strip(),
        status="ACTIVE",
    )
    db.add(artifact)
    db.flush()
    _save_uploaded_file(db, project_id, artifact.id, file.filename or title, raw)
    replaced_count = _soft_delete_matching_uploads(db, project_id, artifact.id, file.filename or title)
    _clear_artifact_chunks(db, artifact_id=artifact.id)
    db.commit()
    db.refresh(artifact)
    return {**_artifact_response(artifact), "parse_status": status, "replaced_count": replaced_count}


@app.post("/api/artifacts/reparse-stored-files")
def reparse_stored_files(
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("artifact:manage")),
) -> dict:
    return _reparse_stored_artifacts(db)


@app.post("/api/projects/{project_id}/artifacts/batch")
def create_artifacts_batch(
    project_id: int,
    payload: ArtifactBatchCreate,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("artifact:manage")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    require_project_access(db, project_id, current_user)
    if not payload.items:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "批量资料不能为空"})
    artifacts = []
    for item in payload.items:
        if item.artifact_type not in ARTIFACT_TYPES:
            raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "资料类型不支持"})
        validate_project_item(db, project_id, item.project_item_id)
        artifact = models.ProjectArtifact(project_id=project_id, **item.model_dump())
        db.add(artifact)
        db.flush()
        _clear_artifact_chunks(db, artifact_id=artifact.id)
        artifacts.append(artifact)
    db.commit()
    for artifact in artifacts:
        db.refresh(artifact)
    return {
        "count": len(artifacts),
        "items": [
            _artifact_response(artifact)
            for artifact in artifacts
        ],
    }


@app.get("/api/projects/{project_id}/artifacts")
def list_artifacts(project_id: int, project_item_id: int | None = None, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    require_project_access(db, project_id, current_user)
    query = db.query(models.ProjectArtifact).filter_by(project_id=project_id, status="ACTIVE")
    if project_item_id is not None:
        query = query.filter_by(project_item_id=project_item_id)
    artifacts = query.order_by(models.ProjectArtifact.id.desc()).all()
    return [_artifact_list_payload(artifact, db) for artifact in artifacts]


@app.get("/api/artifacts/query")
def query_artifacts(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    projects = {project.id: project for project in visible_project_query(db, current_user).all()}
    managed_map = {project.name: _project_management_row(project) for project in projects.values()}
    artifacts = db.query(models.ProjectArtifact).filter(models.ProjectArtifact.status == "ACTIVE", models.ProjectArtifact.project_id.in_(list(projects.keys()))).order_by(models.ProjectArtifact.project_id, models.ProjectArtifact.id.desc()).all()
    return [
        {
            "project_id": artifact.project_id,
            "project_name": projects.get(artifact.project_id).name if projects.get(artifact.project_id) else f"项目 {artifact.project_id}",
            "project_code": projects.get(artifact.project_id).code if projects.get(artifact.project_id) else "",
            "project_manager": managed_map.get(projects.get(artifact.project_id).name, {}).get("project_manager", "未填") if projects.get(artifact.project_id) else "未填",
            "project_intro": managed_map.get(projects.get(artifact.project_id).name, {}).get("technical_terms") if projects.get(artifact.project_id) and managed_map.get(projects.get(artifact.project_id).name, {}).get("technical_terms") else (projects.get(artifact.project_id).description if projects.get(artifact.project_id) else ""),
            **_artifact_list_payload(artifact, db),
        }
        for artifact in artifacts
    ]


@app.get("/api/artifacts/{artifact_id}/file")
def get_artifact_file(
    artifact_id: int,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("read")),
    current_user: dict = Depends(get_current_user),
):
    artifact = db.get(models.ProjectArtifact, artifact_id)
    if artifact is None or artifact.status != "ACTIVE":
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "资料不存在"})
    require_project_access(db, artifact.project_id, current_user)
    filename = _extract_uploaded_filename(artifact.content)
    if not filename:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "资料未关联原始附件"})
    path = _find_stored_upload(artifact.project_id, artifact.id, filename, db)
    if path is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "原始附件不存在"})
    content_type = _extract_uploaded_content_type(artifact.content, filename)
    return FileResponse(path, media_type=content_type, filename=filename, headers={"Content-Disposition": f'inline; filename="{_safe_upload_filename(filename)}"'})


@app.get("/api/items/{item_id}/nodes")
def list_nodes(item_id: int, db: Session = Depends(get_db)) -> list[dict]:
    nodes = db.query(models.CalcNode).filter_by(project_item_id=item_id).order_by(models.CalcNode.sort_order).all()
    return [
        {
            "id": n.id,
            "parent_id": n.parent_id,
            "name": n.name,
            "node_type": n.node_type,
            "template_id": n.template_id,
            "sort_order": n.sort_order,
        }
        for n in nodes
    ]


@app.get("/api/templates")
def list_templates(db: Session = Depends(get_db)) -> list[dict]:
    templates = db.query(models.CalcStepTemplate).all()
    return [{"id": t.id, "code": t.code, "name": t.name, "furnace_type": t.furnace_type} for t in templates]


@app.post("/api/templates")
def create_template(
    payload: TemplateCreate,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("template:manage")),
) -> dict:
    template = models.CalcStepTemplate(
        code=payload.code,
        name=payload.name,
        category=payload.category,
        step_type=payload.step_type,
        furnace_type=payload.furnace_type,
        version=payload.version,
        executor_type=payload.executor_type,
        entrypoint=payload.entrypoint,
        input_fields_json=json.dumps([i.model_dump() for i in payload.input_fields], ensure_ascii=False),
        output_fields_json=json.dumps([o.model_dump() for o in payload.output_fields], ensure_ascii=False),
        report_template_code=payload.report_template_code,
        workflow_type=payload.workflow_type,
        formula_source=payload.formula_source,
        applicable_scope=payload.applicable_scope,
        status=payload.status,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return {"id": template.id, "code": template.code}


@app.post("/api/nodes/{node_id}/executions", response_model=ExecutorResponse)
def execute_node(
    node_id: int,
    payload: ExecutionRequest,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("execution:run")),
    current_user: dict = Depends(get_current_user),
) -> ExecutorResponse:
    node = db.get(models.CalcNode, node_id)
    if node is None or node.template_id is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "计算节点不存在或未绑定模板"})
    item = db.get(models.ProjectItem, node.project_item_id)
    template = db.get(models.CalcStepTemplate, node.template_id)
    if item is None or template is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "名目或模板不存在"})
    require_project_access(db, item.project_id, current_user)

    execution_no = f"EXEC-{utc_now().strftime('%Y%m%d%H%M%S%f')}-{node_id}-{uuid4().hex[:8]}"
    started = utc_now()
    begin = perf_counter()
    execution = models.CalcExecution(
        execution_no=execution_no,
        project_id=item.project_id,
        project_item_id=item.id,
        node_id=node.id,
        template_id=template.id,
        status="RUNNING",
        input_snapshot_json=json.dumps(payload.inputs, ensure_ascii=False),
        template_snapshot_json=json.dumps({"code": template.code, "version": template.version}, ensure_ascii=False),
        executor_version="v1.0-mock",
        started_at=started,
    )
    db.add(execution)
    db.flush()

    response = run_template(template.code, payload.inputs)
    execution.status = "SUCCESS" if response.success else "FAILED"
    execution.finished_at = utc_now()
    execution.duration_ms = int((perf_counter() - begin) * 1000)
    db.add(
        models.CalcResult(
            execution_id=execution.id,
            success=response.success,
            feasible=response.feasible,
            output_json=response.model_dump_json(),
            warnings_json=json.dumps(response.warnings, ensure_ascii=False),
            errors_json=json.dumps(response.errors, ensure_ascii=False),
            logs_json=json.dumps(response.logs, ensure_ascii=False),
        )
    )
    _create_draft_report(db, execution, str(current_user.get("name") or "系统"))
    db.commit()
    return response


@app.get("/api/executions/{execution_id}")
def get_execution(execution_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    execution = db.get(models.CalcExecution, execution_id)
    if execution is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "执行记录不存在"})
    require_project_access(db, execution.project_id, current_user)
    result = execution.result
    return {
        "id": execution.id,
        "execution_no": execution.execution_no,
        "status": execution.status,
        "duration_ms": execution.duration_ms,
        "result": json.loads(result.output_json) if result else None,
    }


@app.get("/api/executions")
def list_executions(limit: int = 20, offset: int = 0, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    safe_limit = max(1, min(limit, 100))
    safe_offset = max(offset, 0)
    executions = (
        db.query(models.CalcExecution)
        .options(joinedload(models.CalcExecution.result))
        .filter(models.CalcExecution.project_id.in_([project.id for project in visible_project_query(db, current_user).all()]))
        .order_by(models.CalcExecution.id.desc())
        .offset(safe_offset)
        .limit(safe_limit)
        .all()
    )
    rows = []
    for execution in executions:
        result = execution.result
        rows.append(
            {
                "id": execution.id,
                "execution_no": execution.execution_no,
                "project_id": execution.project_id,
                "project_item_id": execution.project_item_id,
                "node_id": execution.node_id,
                "status": execution.status,
                "duration_ms": execution.duration_ms,
                "result_id": result.id if result else None,
                "feasible": result.feasible if result else None,
            }
        )
    return rows


def _approval_step_payload(step: models.ApprovalStep) -> dict:
    return {
        "id": step.id,
        "step_order": step.step_order,
        "role_code": step.role_code,
        "role_name": ROLE_NAMES.get(step.role_code, step.role_code),
        "approver_user_id": step.approver_user_id,
        "approver_name": step.approver_name,
        "status": step.status,
        "comment": step.comment,
        "acted_at": step.acted_at.isoformat() if step.acted_at else None,
    }


def _approval_payload(db: Session, approval: models.ApprovalRequest) -> dict:
    steps = (
        db.query(models.ApprovalStep)
        .filter_by(approval_request_id=approval.id)
        .order_by(models.ApprovalStep.step_order.asc())
        .all()
    )
    execution = db.get(models.CalcExecution, approval.execution_id)
    submitter = USER_BY_ID.get(approval.submitted_by)
    current_approver = USER_BY_ID.get(approval.current_approver_id) if approval.current_approver_id else None
    return {
        "id": approval.id,
        "execution_id": approval.execution_id,
        "execution_no": execution.execution_no if execution else None,
        "project_id": execution.project_id if execution else None,
        "status": approval.status,
        "submitted_by": approval.submitted_by,
        "submitted_by_name": submitter["name"] if submitter else str(approval.submitted_by),
        "submitted_at": approval.submitted_at.isoformat() if approval.submitted_at else None,
        "current_approver_id": approval.current_approver_id,
        "current_approver_name": current_approver["name"] if current_approver else None,
        "steps": [_approval_step_payload(step) for step in steps],
    }


def _report_payload(report: models.GeneratedReport) -> dict:
    return {
        "id": report.id,
        "report_no": report.report_no,
        "execution_id": report.execution_id,
        "status": report.status,
        "version": report.version,
        "file_path": report.file_path,
        "watermark": report.watermark,
    }


def _report_version(db: Session, execution_id: int) -> str:
    count = db.query(models.GeneratedReport).filter_by(execution_id=execution_id).count()
    return f"1.{count}"


def _build_report_file_path(execution: models.CalcExecution, version: str) -> str:
    return f"storage/projects/{execution.project_id}/reports/{execution.execution_no}-v{version}.txt"


def _create_draft_report(db: Session, execution: models.CalcExecution, creator_name: str) -> models.GeneratedReport:
    version = _report_version(db, execution.id)
    report = models.GeneratedReport(
        report_no=None,
        execution_id=execution.id,
        status="DRAFT",
        version=version,
        file_path=_build_report_file_path(execution, version),
        watermark=f"草稿 / {creator_name}",
    )
    db.add(report)
    return report


def _get_current_approval_step(db: Session, approval_id: int) -> models.ApprovalStep | None:
    return (
        db.query(models.ApprovalStep)
        .filter_by(approval_request_id=approval_id, status="PENDING")
        .order_by(models.ApprovalStep.step_order.asc())
        .first()
    )


def _get_next_waiting_step(db: Session, approval_id: int, current_order: int) -> models.ApprovalStep | None:
    return (
        db.query(models.ApprovalStep)
        .filter(
            models.ApprovalStep.approval_request_id == approval_id,
            models.ApprovalStep.step_order > current_order,
            models.ApprovalStep.status == "WAITING",
        )
        .order_by(models.ApprovalStep.step_order.asc())
        .first()
    )


@app.post("/api/executions/{execution_id}/approval")
def submit_approval(
    execution_id: int,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("approval:submit")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    execution = db.get(models.CalcExecution, execution_id)
    if execution is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "执行记录不存在"})
    require_project_access(db, execution.project_id, current_user)
    latest_approval = (
        db.query(models.ApprovalRequest)
        .filter(models.ApprovalRequest.execution_id == execution_id)
        .order_by(models.ApprovalRequest.id.desc())
        .first()
    )
    active_approval = (
        db.query(models.ApprovalRequest)
        .filter(
            models.ApprovalRequest.execution_id == execution_id,
            models.ApprovalRequest.status.in_(["SUBMITTED", "IN_REVIEW"]),
        )
        .order_by(models.ApprovalRequest.id.desc())
        .first()
    )
    if active_approval is not None:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "当前执行已有审批流程在进行中"})
    if latest_approval is not None and latest_approval.status == "RETURNED":
        latest_approval.status = "IN_REVIEW"
        latest_approval.submitted_by = int(current_user["id"])
        latest_approval.submitted_at = utc_now()
        latest_approval.current_approver_id = APPROVAL_FLOW_USER_IDS[0]
        steps = (
            db.query(models.ApprovalStep)
            .filter(models.ApprovalStep.approval_request_id == latest_approval.id)
            .order_by(models.ApprovalStep.step_order.asc())
            .all()
        )
        for index, step in enumerate(steps, start=1):
            step.status = "PENDING" if index == 1 else "WAITING"
            step.comment = None
            step.acted_at = None
        db.add(
            models.ApprovalLog(
                approval_request_id=latest_approval.id,
                action="resubmit",
                from_status="RETURNED",
                to_status="IN_REVIEW",
                actor_user_id=int(current_user["id"]),
            )
        )
        db.commit()
        db.refresh(latest_approval)
        return _approval_payload(db, latest_approval)
    approval = models.ApprovalRequest(
        execution_id=execution.id,
        status="IN_REVIEW",
        submitted_by=int(current_user["id"]),
        submitted_at=utc_now(),
        current_approver_id=APPROVAL_FLOW_USER_IDS[0],
    )
    db.add(approval)
    db.flush()
    for index, approver_id in enumerate(APPROVAL_FLOW_USER_IDS, start=1):
        approver = USER_BY_ID[approver_id]
        db.add(
            models.ApprovalStep(
                approval_request_id=approval.id,
                step_order=index,
                role_code=approver["role"],
                approver_user_id=approver_id,
                approver_name=approver["name"],
                status="PENDING" if index == 1 else "WAITING",
            )
        )
    db.add(
        models.ApprovalLog(
            approval_request_id=approval.id,
            action="submit",
            from_status="DRAFT",
            to_status="IN_REVIEW",
            actor_user_id=int(current_user["id"]),
        )
    )
    db.commit()
    db.refresh(approval)
    return _approval_payload(db, approval)


@app.get("/api/approvals")
def list_approvals(
    status: str | None = None,
    mine_only: bool = False,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("read")),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    query = db.query(models.ApprovalRequest).order_by(models.ApprovalRequest.id.desc())
    if status:
        query = query.filter_by(status=status)
    visible_project_ids = {project.id for project in visible_project_query(db, current_user).all()}
    rows = []
    for row in query.all():
        execution = db.get(models.CalcExecution, row.execution_id)
        if execution and execution.project_id in visible_project_ids:
            rows.append(row)
    if mine_only:
        user_id = int(current_user["id"])
        rows = [row for row in rows if row.current_approver_id == user_id or row.submitted_by == user_id]
    return [_approval_payload(db, approval) for approval in rows]


@app.get("/api/approvals/{approval_id}")
def get_approval(approval_id: int, db: Session = Depends(get_db), role: str = Depends(permission_dependency("read")), current_user: dict = Depends(get_current_user)) -> dict:
    approval = db.get(models.ApprovalRequest, approval_id)
    if approval is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "审批申请不存在"})
    execution = db.get(models.CalcExecution, approval.execution_id)
    if execution is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "执行记录不存在"})
    require_project_access(db, execution.project_id, current_user)
    return _approval_payload(db, approval)


@app.post("/api/approvals/{approval_id}/approve")
def approve_request(
    approval_id: int,
    payload: ApprovalActionRequest | None = None,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("approval:review")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    approval = db.get(models.ApprovalRequest, approval_id)
    if approval is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "审批申请不存在"})
    execution = db.get(models.CalcExecution, approval.execution_id)
    if execution is not None:
        require_project_access(db, execution.project_id, current_user)
    if approval.status != "IN_REVIEW":
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "当前状态不可审批通过"})
    current_step = _get_current_approval_step(db, approval.id)
    if current_step is None:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "当前审批没有待处理步骤"})
    if current_step.approver_user_id != int(current_user["id"]):
        raise HTTPException(status_code=403, detail={"code": "PERMISSION_DENIED", "message": "当前用户不是该步骤审批人"})
    current_step.status = "APPROVED"
    current_step.comment = payload.comment if payload else None
    current_step.acted_at = utc_now()
    next_step = _get_next_waiting_step(db, approval.id, current_step.step_order)
    from_status = approval.status
    if next_step is None:
        approval.status = "APPROVED"
        approval.current_approver_id = None
        action = "approve_final"
    else:
        next_step.status = "PENDING"
        approval.current_approver_id = next_step.approver_user_id
        action = "approve_step"
    db.add(
        models.ApprovalLog(
            approval_request_id=approval.id,
            action=action,
            from_status=from_status,
            to_status=approval.status,
            comment=payload.comment if payload else None,
            actor_user_id=int(current_user["id"]),
        )
    )
    db.commit()
    db.refresh(approval)
    return _approval_payload(db, approval)


@app.post("/api/approvals/{approval_id}/return")
def return_request(
    approval_id: int,
    payload: ApprovalActionRequest | None = None,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("approval:review")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    approval = db.get(models.ApprovalRequest, approval_id)
    if approval is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "审批申请不存在"})
    execution = db.get(models.CalcExecution, approval.execution_id)
    if execution is not None:
        require_project_access(db, execution.project_id, current_user)
    if approval.status != "IN_REVIEW":
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "当前状态不可退回"})
    current_step = _get_current_approval_step(db, approval.id)
    if current_step is None:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "当前审批没有待处理步骤"})
    if current_step.approver_user_id != int(current_user["id"]):
        raise HTTPException(status_code=403, detail={"code": "PERMISSION_DENIED", "message": "当前用户不是该步骤审批人"})
    current_step.status = "RETURNED"
    current_step.comment = payload.comment if payload else None
    current_step.acted_at = utc_now()
    approval.status = "RETURNED"
    approval.current_approver_id = approval.submitted_by
    db.add(
        models.ApprovalLog(
            approval_request_id=approval.id,
            action="return",
            from_status="IN_REVIEW",
            to_status="RETURNED",
            comment=payload.comment if payload else None,
            actor_user_id=int(current_user["id"]),
        )
    )
    db.commit()
    db.refresh(approval)
    return _approval_payload(db, approval)


@app.post("/api/executions/{execution_id}/reports")
def create_report(
    execution_id: int,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("report:create")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    execution = db.get(models.CalcExecution, execution_id)
    if execution is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "执行记录不存在"})
    require_project_access(db, execution.project_id, current_user)
    report = _create_draft_report(db, execution, str(current_user.get("name") or "系统"))
    db.commit()
    db.refresh(report)
    return _report_payload(report)


@app.get("/api/reports")
def list_reports(
    execution_id: int | None = None,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("read")),
    current_user: dict = Depends(get_current_user),
) -> list[dict]:
    query = db.query(models.GeneratedReport).order_by(models.GeneratedReport.id.desc())
    if execution_id is not None:
        query = query.filter_by(execution_id=execution_id)
    visible_project_ids = {project.id for project in visible_project_query(db, current_user).all()}
    reports = []
    for report in query.all():
        execution = db.get(models.CalcExecution, report.execution_id)
        if execution and execution.project_id in visible_project_ids:
            reports.append(report)
    return [_report_payload(report) for report in reports]


@app.post("/api/reports/{report_id}/publish")
def publish_report(
    report_id: int,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("report:publish")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    report = db.get(models.GeneratedReport, report_id)
    if report is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "报告不存在"})
    execution = db.get(models.CalcExecution, report.execution_id)
    if execution is not None:
        require_project_access(db, execution.project_id, current_user)
    if report.status != "DRAFT":
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "当前报告状态不可发布"})
    approved = (
        db.query(models.ApprovalRequest)
        .filter_by(execution_id=report.execution_id, status="APPROVED")
        .order_by(models.ApprovalRequest.id.desc())
        .first()
    )
    if approved is None:
        raise HTTPException(status_code=409, detail={"code": "STATE_INVALID", "message": "审批完成后才能发布正式报告"})
    report.status = "OFFICIAL"
    report.report_no = f"RPT-{utc_now().strftime('%Y%m%d')}-{uuid4().hex[:6]}"
    report.watermark = None
    db.commit()
    db.refresh(report)
    return {**_report_payload(report), "published_by": int(current_user["id"])}


@app.get("/api/reports/{report_id}")
def get_report(report_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> dict:
    report = db.get(models.GeneratedReport, report_id)
    if report is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "报告不存在"})
    execution = db.get(models.CalcExecution, report.execution_id)
    if execution is not None:
        require_project_access(db, execution.project_id, current_user)
    return _report_payload(report)


@app.get("/api/reports/{report_id}/download", response_class=PlainTextResponse)
def download_report(
    report_id: int,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("report:download")),
    current_user: dict = Depends(get_current_user),
) -> str:
    report = db.get(models.GeneratedReport, report_id)
    if report is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "报告不存在"})
    execution = db.get(models.CalcExecution, report.execution_id)
    if execution is not None:
        require_project_access(db, execution.project_id, current_user)
    result = db.query(models.CalcResult).filter_by(execution_id=report.execution_id).one_or_none()
    approval = (
        db.query(models.ApprovalRequest)
        .filter_by(execution_id=report.execution_id)
        .order_by(models.ApprovalRequest.id.desc())
        .first()
    )
    output = json.loads(result.output_json) if result else {}
    return "\n".join(
        [
            "工业炉计算报告 V1.0",
            f"报告状态: {report.status}",
            f"报告编号: {report.report_no or 'DRAFT'}",
            f"报告版本: {report.version}",
            f"执行编号: {execution.execution_no if execution else ''}",
            f"审批状态: {approval.status if approval else '未提交'}",
            f"是否可行: {output.get('feasible')}",
            f"输出结果: {json.dumps(output.get('outputs', {}), ensure_ascii=False)}",
            f"水印: {report.watermark or '无'}",
        ]
    )


@app.post("/api/comparisons")
def create_comparison(
    payload: dict,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("comparison:create")),
) -> dict:
    step_type = payload.get("step_type")
    name = payload.get("name") or f"{step_type} 对比组"
    result_ids = payload.get("result_ids", [])
    if not step_type or not result_ids:
        raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": "step_type 和 result_ids 必填"})
    group = models.ComparisonGroup(name=name, step_type=step_type, created_by=2)
    db.add(group)
    db.flush()
    for result_id in result_ids:
        if db.get(models.CalcResult, result_id) is None:
            raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": f"结果不存在: {result_id}"})
        db.add(models.ComparisonItem(comparison_group_id=group.id, result_id=result_id))
    db.commit()
    return {"id": group.id, "name": group.name, "step_type": group.step_type}


@app.get("/api/comparisons/{group_id}")
def get_comparison(group_id: int, db: Session = Depends(get_db)) -> dict:
    group = db.get(models.ComparisonGroup, group_id)
    if group is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "对比组不存在"})
    items = db.query(models.ComparisonItem).filter_by(comparison_group_id=group.id).all()
    results = []
    for item in items:
        result = db.get(models.CalcResult, item.result_id)
        if result:
            payload = json.loads(result.output_json)
            results.append({"result_id": result.id, "outputs": payload.get("outputs", {}), "feasible": result.feasible})
    return {"id": group.id, "name": group.name, "step_type": group.step_type, "results": results}


@app.post("/api/projects/{project_id}/ai-analyses")
async def create_ai_analysis(
    project_id: int,
    payload: AiAnalysisRequest,
    db: Session = Depends(get_db),
    role: str = Depends(permission_dependency("ai:analyze")),
    current_user: dict = Depends(get_current_user),
) -> dict:
    require_project_access(db, project_id, current_user)
    validate_project_item(db, project_id, payload.project_item_id)
    executions = []
    for execution_id in payload.execution_ids:
        execution = db.get(models.CalcExecution, execution_id)
        if execution is None:
            raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": f"执行记录不存在: {execution_id}"})
        if execution.project_id != project_id:
            raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": f"执行记录不属于当前项目: {execution_id}"})
        result = db.query(models.CalcResult).filter_by(execution_id=execution.id).one_or_none()
        executions.append(
            {
                "execution_id": execution.id,
                "execution_no": execution.execution_no,
                "inputs": json.loads(execution.input_snapshot_json),
                "result": json.loads(result.output_json) if result else None,
            }
        )
    selected_artifact_ids = payload.artifact_ids or [artifact.id for artifact in db.query(models.ProjectArtifact).filter_by(project_id=project_id, status="ACTIVE").all()]
    artifacts = []
    for artifact_id in selected_artifact_ids:
        artifact = db.get(models.ProjectArtifact, artifact_id)
        if artifact is None:
            raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": f"资料不存在: {artifact_id}"})
        if artifact.project_id != project_id:
            raise HTTPException(status_code=400, detail={"code": "PARAM_INVALID", "message": f"资料不属于当前项目: {artifact_id}"})
        artifacts.append(
            {
                "artifact_id": artifact.id,
                "type": artifact.artifact_type,
                "type_name": ARTIFACT_TYPES.get(artifact.artifact_type, artifact.artifact_type),
                "title": artifact.title,
                "content_preview": _content_preview(artifact.content),
                "content": artifact.content,
            }
        )
    pasted_images = _build_pasted_image_context(payload.pasted_images)
    ai_question = payload.question.strip()
    if pasted_images:
        image_lines = ["补充图片信息:"]
        for index, image in enumerate(pasted_images, start=1):
            image_lines.append(f"- 图片{index}: {image['name']} | {image['content_type']} | {image['parse_status']}")
            if image["summary"].strip():
                image_lines.append(image["summary"])
        ai_question = "\n\n".join([ai_question or "请结合项目资料和粘贴图片回答问题。", "\n".join(image_lines)])
    retrieved_artifacts = await _search_project_artifacts_for_ai(db, project_id, ai_question or payload.question, selected_artifact_ids)
    request_data = {
        "equipment_name": payload.equipment_name,
        "analysis_type": payload.analysis_type,
        "question": ai_question,
        "original_question": payload.question,
        "pasted_images": pasted_images,
        "executions": executions,
        "retrieved_artifacts": retrieved_artifacts,
        "artifacts": artifacts,
    }
    prompt = json.dumps(request_data, ensure_ascii=False, indent=2)
    ai_result = await run_joint_analysis(prompt)
    analysis = models.AiAnalysis(
        project_id=project_id,
        project_item_id=payload.project_item_id,
        equipment_name=payload.equipment_name,
        analysis_type=payload.analysis_type,
        request_json=json.dumps(request_data, ensure_ascii=False),
        response_json=json.dumps(ai_result, ensure_ascii=False),
        provider=ai_result.get("provider", "unknown"),
        status="SUCCESS" if "errors" not in ai_result else "FAILED",
    )
    db.add(analysis)
    db.commit()
    db.refresh(analysis)
    return {"id": analysis.id, "provider": analysis.provider, "status": analysis.status, "result": ai_result}


@app.get("/api/projects/{project_id}/ai-analyses")
def list_ai_analyses(project_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)) -> list[dict]:
    require_project_access(db, project_id, current_user)
    analyses = db.query(models.AiAnalysis).filter_by(project_id=project_id).order_by(models.AiAnalysis.id.desc()).all()
    return [
        {
            "id": analysis.id,
            "equipment_name": analysis.equipment_name,
            "analysis_type": analysis.analysis_type,
            "provider": analysis.provider,
            "status": analysis.status,
            "result": json.loads(analysis.response_json),
        }
        for analysis in analyses
    ]
